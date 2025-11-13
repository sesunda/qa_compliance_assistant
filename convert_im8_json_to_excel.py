"""
Convert IM8 JSON files to Excel format
Generates proper Excel workbooks from JSON assessment templates
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_template_excel(json_file, output_file):
    """Convert IM8 template JSON to Excel with multiple sheets"""
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    domain_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    domain_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Create Cover Page
    ws_cover = wb.create_sheet("Cover Page")
    cover_data = [
        ["IM8 INFORMATION SECURITY ASSESSMENT", ""],
        ["", ""],
        ["Template Version:", data.get("template_version", "1.0")],
        ["Framework:", data.get("framework", "IM8")],
        ["Created Date:", data.get("created_date", "")],
        ["Created By:", data.get("created_by", "")],
        ["", ""],
        ["Agency Name:", data.get("cover_page", {}).get("agency_name", "[To be filled]")],
        ["Assessment Period:", data.get("cover_page", {}).get("assessment_period", "[To be filled]")],
        ["Prepared By:", data.get("cover_page", {}).get("prepared_by", "[To be filled]")],
        ["Preparation Date:", data.get("cover_page", {}).get("preparation_date", "[To be filled]")],
        ["", ""],
        ["Reviewed By:", data.get("cover_page", {}).get("reviewed_by", "[To be filled]")],
        ["Review Date:", data.get("cover_page", {}).get("review_date", "[To be filled]")],
        ["Approval Status:", data.get("cover_page", {}).get("approval_status", "Pending")],
    ]
    
    for row in cover_data:
        ws_cover.append(row)
    
    # Format cover page
    ws_cover['A1'].font = Font(bold=True, size=16, color="366092")
    ws_cover.merge_cells('A1:B1')
    ws_cover.column_dimensions['A'].width = 25
    ws_cover.column_dimensions['B'].width = 40
    
    # 2. Create Instructions Sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = data.get("instructions", {})
    
    ws_instructions.append(["IM8 ASSESSMENT INSTRUCTIONS"])
    ws_instructions.append([""])
    ws_instructions.append(["Overview:"])
    ws_instructions.append([instructions.get("overview", "")])
    ws_instructions.append([""])
    ws_instructions.append(["Completion Steps:"])
    
    for step in instructions.get("completion_steps", []):
        ws_instructions.append([step])
    
    ws_instructions.append([""])
    ws_instructions.append(["Implementation Status Options:"])
    for status in instructions.get("status_options", []):
        ws_instructions.append([status])
    
    ws_instructions['A1'].font = Font(bold=True, size=14, color="366092")
    ws_instructions.column_dimensions['A'].width = 80
    
    # 3. Create Domain Sheets
    domains = data.get("domains", [])
    
    for domain in domains:
        domain_id = domain.get("domain_id", "")
        domain_name = domain.get("domain_name", "")
        sheet_name = f"{domain_id}"[:31]  # Excel sheet name limit
        
        ws = wb.create_sheet(sheet_name)
        
        # Domain header
        ws.append([f"{domain_id}: {domain_name}"])
        ws.append([domain.get("description", "")])
        ws.append([""])
        
        # Format domain header
        ws['A1'].font = domain_font
        ws['A1'].fill = domain_fill
        ws.merge_cells(f'A1:H1')
        ws['A2'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A2:H2')
        
        # Controls table header
        headers = [
            "Control ID",
            "Control Name",
            "Description",
            "Implementation Status",
            "Evidence Reference",
            "Evidence Files",
            "Notes",
            "Last Review Date"
        ]
        ws.append(headers)
        
        # Format headers
        header_row = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add controls
        controls = domain.get("controls", [])
        for control in controls:
            row_data = [
                control.get("control_id", ""),
                control.get("control_name", ""),
                control.get("description", ""),
                control.get("implementation_status", ""),
                control.get("evidence_reference", ""),
                ", ".join(control.get("evidence_files", [])) if isinstance(control.get("evidence_files"), list) else "",
                control.get("notes", ""),
                control.get("last_review_date", "")
            ]
            ws.append(row_data)
            
            # Format control row
            current_row = ws.max_row
            for col_num in range(1, 9):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 15
    
    # 4. Create Summary Dashboard
    ws_summary = wb.create_sheet("Summary Dashboard", 0)
    
    summary = data.get("summary_dashboard", {})
    summary_data = [
        ["IM8 ASSESSMENT SUMMARY DASHBOARD", ""],
        ["", ""],
        ["Total Controls:", summary.get("total_controls", 0)],
        ["Domains Completed:", summary.get("domains_completed", 0)],
        ["Completion Percentage:", f"{summary.get('completion_percentage', 0)}%"],
        ["Evidence Items Attached:", summary.get("evidence_items_attached", 0)],
        ["", ""],
        ["Controls by Status:", ""],
        ["  Implemented:", summary.get("controls_by_status", {}).get("implemented", 0)],
        ["  Partial:", summary.get("controls_by_status", {}).get("partial", 0)],
        ["  Not Implemented:", summary.get("controls_by_status", {}).get("not_implemented", 0)],
        ["  Not Applicable:", summary.get("controls_by_status", {}).get("not_applicable", 0)],
        ["", ""],
        ["Validation Status:", summary.get("validation_status", "")],
        ["Submitted By:", summary.get("submitted_by", "")],
        ["Submission Date:", summary.get("submission_date", "")],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
    ws_summary.merge_cells('A1:B1')
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 30
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Created: {output_file}")


def create_completed_excel(json_file, output_file):
    """Convert completed IM8 assessment JSON to Excel"""
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    domain_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    domain_font = Font(bold=True, size=12)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Cover Page
    ws_cover = wb.create_sheet("Cover Page")
    cover = data.get("cover_page", {})
    cover_data = [
        ["IM8 INFORMATION SECURITY ASSESSMENT", ""],
        ["SECURITY AWARENESS TRAINING", ""],
        ["", ""],
        ["Agency Name:", cover.get("agency_name", "")],
        ["Assessment Period:", cover.get("assessment_period", "")],
        ["Prepared By:", cover.get("prepared_by", "")],
        ["Preparation Date:", cover.get("preparation_date", "")],
        ["", ""],
        ["Reviewed By:", cover.get("reviewed_by", "")],
        ["Review Date:", cover.get("review_date", "")],
        ["Approval Status:", cover.get("approval_status", "")],
    ]
    
    for row in cover_data:
        ws_cover.append(row)
    
    ws_cover['A1'].font = Font(bold=True, size=16, color="366092")
    ws_cover['A2'].font = Font(bold=True, size=14, color="366092")
    ws_cover.merge_cells('A1:B1')
    ws_cover.merge_cells('A2:B2')
    ws_cover.column_dimensions['A'].width = 25
    ws_cover.column_dimensions['B'].width = 50
    
    # 2. Assessment Summary
    ws_summary = wb.create_sheet("Assessment Summary")
    summary = data.get("assessment_summary", {})
    
    summary_data = [
        ["ASSESSMENT SUMMARY", ""],
        ["", ""],
        ["Focus Domain:", summary.get("focus_domain", "")],
        ["Focus Control:", summary.get("focus_control", "")],
        ["Implementation Status:", summary.get("implementation_status", "")],
        ["Overall Compliance Score:", f"{summary.get('overall_compliance_score', 0)}%"],
        ["Evidence Items:", summary.get("evidence_items", 0)],
        ["", ""],
        ["KEY FINDINGS:", ""],
    ]
    
    for finding in summary.get("key_findings", []):
        summary_data.append([f"✓ {finding}", ""])
    
    summary_data.append(["", ""])
    summary_data.append(["RECOMMENDATIONS:", ""])
    
    for rec in summary.get("recommendations", []):
        summary_data.append([f"→ {rec}", ""])
    
    for row in summary_data:
        ws_summary.append(row)
    
    ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
    ws_summary.merge_cells('A1:B1')
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 60
    
    # 3. Control Details (IM8-01-03)
    domains = data.get("domains", [])
    for domain in domains:
        domain_id = domain.get("domain_id", "")
        ws = wb.create_sheet(domain_id)
        
        # Domain header
        ws.append([f"{domain_id}: {domain.get('domain_name', '')}"])
        ws.append([domain.get("description", "")])
        ws.append([""])
        
        controls = domain.get("controls", [])
        for control in controls:
            # Control header
            ws.append([f"Control: {control.get('control_id', '')}", ""])
            ws.append([control.get('control_name', ''), ""])
            ws.append([control.get('description', ''), ""])
            ws.append(["", ""])
            
            # Basic info
            ws.append(["Implementation Status:", control.get('implementation_status', '')])
            ws.append(["Evidence Reference:", control.get('evidence_reference', '')])
            ws.append(["", ""])
            
            # Evidence files
            ws.append(["EVIDENCE FILES:", ""])
            evidence_files = control.get('evidence_files', [])
            if isinstance(evidence_files, list):
                for idx, ef in enumerate(evidence_files, 1):
                    if isinstance(ef, dict):
                        ws.append([f"{idx}. {ef.get('filename', '')}", ""])
                        ws.append([f"   Type: {ef.get('file_type', '')}", ""])
                        ws.append([f"   Size: {ef.get('file_size_kb', 0)} KB", ""])
                        ws.append([f"   Description: {ef.get('description', '')}", ""])
                        ws.append(["", ""])
            
            # Training metrics (if available)
            impl_details = control.get('implementation_details', {})
            if impl_details:
                ws.append(["TRAINING METRICS:", ""])
                metrics = impl_details.get('training_metrics', {})
                ws.append([f"Total Employees: {metrics.get('total_employees', 0)}", ""])
                ws.append([f"Completed Training: {metrics.get('completed_training', 0)}", ""])
                ws.append([f"Completion Rate: {metrics.get('completion_rate_percent', 0)}%", ""])
                ws.append([f"Average Score: {metrics.get('average_score_percent', 0)}%", ""])
                ws.append(["", ""])
            
            # Validation results
            validation = control.get('validation_results', {})
            if validation:
                ws.append(["VALIDATION RESULTS:", ""])
                ws.append([f"Overall Status: {validation.get('overall_status', '')}", ""])
                ws.append([f"Overall Score: {validation.get('overall_score', 0)}/100", ""])
                ws.append([f"Compliance Status: {validation.get('compliance_status', '')}", ""])
                ws.append(["", ""])
            
            # Notes
            ws.append(["NOTES:", ""])
            ws.append([control.get('notes', ''), ""])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70
    
    # Save workbook
    wb.save(output_file)
    print(f"✅ Created: {output_file}")


if __name__ == "__main__":
    # Convert template
    template_json = "storage/IM8_Assessment_Template.json"
    template_excel = "storage/IM8_Assessment_Template.xlsx"
    
    if os.path.exists(template_json):
        print(f"📄 Converting {template_json}...")
        create_template_excel(template_json, template_excel)
    else:
        print(f"❌ File not found: {template_json}")
    
    # Convert completed assessment
    completed_json = "storage/IM8_Security_Awareness_Training_Completed.json"
    completed_excel = "storage/IM8_Security_Awareness_Training_Completed.xlsx"
    
    if os.path.exists(completed_json):
        print(f"📄 Converting {completed_json}...")
        create_completed_excel(completed_json, completed_excel)
    else:
        print(f"❌ File not found: {completed_json}")
    
    print("\n✅ Conversion complete!")
    print(f"   Template: {template_excel}")
    print(f"   Completed: {completed_excel}")
