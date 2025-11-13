#!/usr/bin/env python3
"""
Convert IM8 JSON assessment files to Excel format
Usage: python convert_json_to_excel.py
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_excel_from_json(json_file_path, output_excel_path):
    """Convert IM8 JSON to formatted Excel workbook"""
    
    print(f"📖 Reading JSON file: {json_file_path}")
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    domain_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    domain_font = Font(bold=True, color="FFFFFF", size=12)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create Cover Page
    print("📄 Creating Cover Page...")
    cover_sheet = wb.create_sheet("Cover Page", 0)
    cover_data = data.get('cover_page', {})
    
    cover_sheet['A1'] = cover_data.get('title', 'IM8 Assessment')
    cover_sheet['A1'].font = Font(bold=True, size=16)
    
    cover_info = [
        ['Agency:', cover_data.get('agency_name', '')],
        ['Assessment Period:', cover_data.get('assessment_period', '')],
        ['Prepared By:', cover_data.get('prepared_by', '')],
        ['Preparation Date:', cover_data.get('preparation_date', '')],
        ['Reviewed By:', cover_data.get('reviewed_by', '')],
        ['Review Date:', cover_data.get('review_date', '')],
        ['Approval Status:', cover_data.get('approval_status', '')]
    ]
    
    row = 3
    for label, value in cover_info:
        cover_sheet[f'A{row}'] = label
        cover_sheet[f'A{row}'].font = Font(bold=True)
        cover_sheet[f'B{row}'] = value
        row += 1
    
    # Add instructions if present
    if 'instructions' in data:
        instructions = data['instructions']
        cover_sheet[f'A{row+1}'] = 'Instructions:'
        cover_sheet[f'A{row+1}'].font = Font(bold=True, size=12)
        row += 2
        
        cover_sheet[f'A{row}'] = instructions.get('overview', '')
        row += 2
        
        if 'completion_steps' in instructions:
            cover_sheet[f'A{row}'] = 'Completion Steps:'
            cover_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            for step in instructions['completion_steps']:
                cover_sheet[f'A{row}'] = step
                row += 1
    
    # Adjust column widths
    cover_sheet.column_dimensions['A'].width = 20
    cover_sheet.column_dimensions['B'].width = 50
    
    # Create domain sheets
    domains = data.get('domains', [])
    print(f"📊 Creating {len(domains)} domain sheets...")
    
    for domain in domains:
        domain_id = domain.get('domain_id', 'Unknown')
        domain_name = domain.get('domain_name', 'Unknown')
        sheet_name = f"{domain_id}"[:31]  # Excel sheet name limit
        
        print(f"  → {sheet_name}: {domain_name}")
        ws = wb.create_sheet(sheet_name)
        
        # Domain header
        ws['A1'] = f"{domain_id}: {domain_name}"
        ws['A1'].font = domain_font
        ws['A1'].fill = domain_fill
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:H1')
        
        ws['A2'] = domain.get('description', '')
        ws['A2'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A2:H2')
        ws.row_dimensions[2].height = 30
        
        # Control table headers
        headers = ['Control ID', 'Control Name', 'Description', 'Implementation Status', 
                   'Evidence Reference', 'Evidence Files', 'Notes', 'Last Review']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border
        
        # Add controls
        controls = domain.get('controls', [])
        row = 5
        
        for control in controls:
            ws.cell(row=row, column=1).value = control.get('control_id', '')
            ws.cell(row=row, column=2).value = control.get('control_name', '')
            ws.cell(row=row, column=3).value = control.get('description', '')
            ws.cell(row=row, column=4).value = control.get('implementation_status', '')
            ws.cell(row=row, column=5).value = control.get('evidence_reference', '')
            
            # Evidence files
            evidence_files = control.get('evidence_files', [])
            if evidence_files:
                files_text = '\n'.join([f"• {ef.get('filename', ef) if isinstance(ef, dict) else ef}" 
                                       for ef in evidence_files])
                ws.cell(row=row, column=6).value = files_text
            else:
                ws.cell(row=row, column=6).value = ''
            
            ws.cell(row=row, column=7).value = control.get('notes', '')
            ws.cell(row=row, column=8).value = control.get('last_review_date', '')
            
            # Apply borders and alignment
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = cell_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            ws.row_dimensions[row].height = 60
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 12
    
    # Create Summary Dashboard
    print("📈 Creating Summary Dashboard...")
    summary_sheet = wb.create_sheet("Summary Dashboard")
    summary_data = data.get('summary_dashboard', {})
    
    summary_sheet['A1'] = 'IM8 Assessment Summary Dashboard'
    summary_sheet['A1'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:B1')
    
    summary_items = [
        ['Total Controls:', summary_data.get('total_controls', 0)],
        ['Completion Percentage:', f"{summary_data.get('completion_percentage', 0)}%"],
        ['Domains Completed:', summary_data.get('domains_completed', 0)],
        ['Evidence Items Attached:', summary_data.get('evidence_items_attached', 0)],
        ['Validation Status:', summary_data.get('validation_status', '')],
        ['Last Updated:', summary_data.get('last_updated', '')],
        ['Submitted By:', summary_data.get('submitted_by', '')],
        ['Submission Date:', summary_data.get('submission_date', '')]
    ]
    
    row = 3
    for label, value in summary_items:
        summary_sheet[f'A{row}'] = label
        summary_sheet[f'A{row}'].font = Font(bold=True)
        summary_sheet[f'B{row}'] = value
        row += 1
    
    # Controls by status
    if 'controls_by_status' in summary_data:
        row += 1
        summary_sheet[f'A{row}'] = 'Controls by Status:'
        summary_sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for status, count in summary_data['controls_by_status'].items():
            summary_sheet[f'A{row}'] = status.replace('_', ' ').title()
            summary_sheet[f'B{row}'] = count
            row += 1
    
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 30
    
    # Save workbook
    print(f"💾 Saving Excel file: {output_excel_path}")
    wb.save(output_excel_path)
    print(f"✅ Excel file created successfully!")
    return output_excel_path


def main():
    """Main conversion function"""
    print("=" * 60)
    print("IM8 JSON to Excel Converter")
    print("=" * 60)
    print()
    
    # Check if openpyxl is installed
    try:
        import openpyxl
    except ImportError:
        print("❌ Error: openpyxl library not found!")
        print("📦 Installing openpyxl...")
        import subprocess
        subprocess.check_call(['pip', 'install', 'openpyxl'])
        print("✅ openpyxl installed successfully!")
        print()
    
    # File paths
    template_json = 'storage/IM8_Assessment_Template.json'
    completed_json = 'storage/IM8_Security_Awareness_Training_Completed.json'
    
    template_excel = 'storage/IM8_Assessment_Template.xlsx'
    completed_excel = 'storage/IM8_Security_Awareness_Training_Completed.xlsx'
    
    # Convert template
    if os.path.exists(template_json):
        print("🔄 Converting Template JSON to Excel...")
        create_excel_from_json(template_json, template_excel)
        print()
    else:
        print(f"⚠️  Template JSON not found: {template_json}")
        print()
    
    # Convert completed assessment
    if os.path.exists(completed_json):
        print("🔄 Converting Completed Assessment JSON to Excel...")
        create_excel_from_json(completed_json, completed_excel)
        print()
    else:
        print(f"⚠️  Completed JSON not found: {completed_json}")
        print()
    
    print("=" * 60)
    print("✅ Conversion Complete!")
    print("=" * 60)
    print()
    print("📁 Generated Files:")
    if os.path.exists(template_excel):
        print(f"  1. {template_excel}")
    if os.path.exists(completed_excel):
        print(f"  2. {completed_excel}")
    print()
    print("📝 Next Steps:")
    print("  1. Open the Excel files to review the structure")
    print("  2. Template file can be shared with Analysts")
    print("  3. Completed file shows example submission with data")
    print()


if __name__ == '__main__':
    main()
