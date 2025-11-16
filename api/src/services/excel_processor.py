"""
IM8 Excel Processor Service
Handles parsing IM8 assessment Excel files and extracting embedded PDFs
"""

from typing import Dict, List, Any, Optional
import io
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class IM8ExcelProcessor:
    """Process IM8 assessment Excel documents"""
    
    # Expected sheet names
    REQUIRED_SHEETS = [
        "Metadata",
        "Domain_1_Info_Security_Governance",
        "Domain_2_Network_Security",
        "Summary",
        "Reference_Policies"
    ]
    
    # Optional sheets
    OPTIONAL_SHEETS = ["Instructions"]
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for IM8 Excel processing. "
                "Install with: pip install openpyxl"
            )
    
    def parse_im8_document(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse IM8 Excel document and extract structured data
        
        Args:
            file_content: Binary content of Excel file
            filename: Original filename
            
        Returns:
            Dictionary with parsed IM8 data structure
            
        Raises:
            ValueError: If file format is invalid
        """
        try:
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {str(e)}")
        
        # Validate required sheets exist
        missing_sheets = [
            sheet for sheet in self.REQUIRED_SHEETS 
            if sheet not in workbook.sheetnames
        ]
        if missing_sheets:
            raise ValueError(
                f"Missing required sheets: {', '.join(missing_sheets)}"
            )
        
        # Parse each section
        metadata = self._parse_metadata_sheet(workbook["Metadata"])
        domain_1 = self._parse_domain_sheet(
            workbook["Domain_1_Info_Security_Governance"],
            "Domain 1: Information Security Governance"
        )
        domain_2 = self._parse_domain_sheet(
            workbook["Domain_2_Network_Security"],
            "Domain 2: Network Security"
        )
        reference_policies = self._parse_reference_policies_sheet(
            workbook["Reference_Policies"]
        )
        summary = self._parse_summary_sheet(workbook["Summary"])
        
        # Combine into structured format
        result = {
            "evidence_type": "im8_assessment_document",
            "framework": "IM8",
            "filename": filename,
            "parsed_at": now_sgt().isoformat(),
            "metadata": metadata,
            "domains": [domain_1, domain_2],
            "reference_policies": reference_policies,
            "summary": summary,
            "validation_status": "pending"
        }
        
        return result
    
    def _parse_metadata_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """Parse Metadata sheet (key-value pairs)"""
        metadata = {}
        
        # Expect format: Column A = Field, Column B = Value
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] and len(row) > 1:  # Has field name
                field = str(row[0]).strip()
                value = row[1] if row[1] is not None else ""
                
                # Convert to snake_case for consistency
                field_key = field.lower().replace(" ", "_")
                metadata[field_key] = value
        
        return metadata
    
    def _parse_domain_sheet(self, sheet: Worksheet, domain_name: str) -> Dict[str, Any]:
        """Parse domain control sheet"""
        controls = []
        
        # Get headers (row 1)
        headers = [cell.value for cell in sheet[1]]
        
        # Expected columns: Control ID, Control Name, Description, Status, Evidence, Implementation Date, Notes
        # Parse data rows (starting from row 2)
        for row in sheet.iter_rows(min_row=2, values_only=False):
            control_id = row[0].value
            
            # Skip empty rows
            if not control_id:
                continue
            
            control = {
                "control_id": str(control_id).strip() if control_id else "",
                "control_name": str(row[1].value).strip() if row[1].value else "",
                "description": str(row[2].value).strip() if row[2].value else "",
                "status": str(row[3].value).strip() if row[3].value else "",
                "implementation_date": self._parse_date(row[5].value) if row[5].value else None,
                "notes": str(row[6].value).strip() if row[6].value else "",
                "has_embedded_evidence": self._check_embedded_file(row[4])
            }
            
            controls.append(control)
        
        return {
            "domain_name": domain_name,
            "domain_id": domain_name.split(":")[0].strip(),  # "Domain 1" or "Domain 2"
            "controls": controls,
            "control_count": len(controls)
        }
    
    def _parse_reference_policies_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Reference_Policies sheet"""
        policies = []
        
        # Parse data rows (starting from row 2)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            policy = {
                "policy_name": str(row[0]).strip() if row[0] else "",
                "version": str(row[1]).strip() if row[1] else "",
                "approval_date": self._parse_date(row[2]) if row[2] else None,
                "document_location": str(row[3]).strip() if row[3] else "",
                "notes": str(row[4]).strip() if row[4] else ""
            }
            
            policies.append(policy)
        
        return policies
    
    def _parse_summary_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """Parse Summary sheet"""
        summary = {}
        
        # Parse key-value pairs (similar to metadata)
        for row in sheet.iter_rows(values_only=True):
            if row[0] and len(row) > 1:
                field = str(row[0]).strip()
                value = row[1]
                
                field_key = field.lower().replace(" ", "_").replace("%", "percentage")
                summary[field_key] = value
        
        return summary
    
    def _check_embedded_file(self, cell) -> bool:
        """
        Check if cell contains embedded file/object
        
        Note: openpyxl has limited support for embedded objects.
        This is a basic check - may need enhancement for production.
        """
        # Check if cell has a hyperlink or object
        if hasattr(cell, 'hyperlink') and cell.hyperlink:
            return True
        
        # Check cell value indicates embedded file
        if cell.value:
            value_str = str(cell.value).lower()
            if 'pdf' in value_str or 'embed' in value_str:
                return True
        
        return False
    
    def _parse_date(self, value: Any) -> Optional[str]:
        """Parse date value to ISO format string"""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.date().isoformat()
        
        # Try parsing string dates
        if isinstance(value, str):
            try:
                # Try common formats
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"]:
                    try:
                        dt = datetime.strptime(value.strip(), fmt)
                        return dt.date().isoformat()
                    except ValueError:
                        continue
            except Exception:
                pass
        
        return str(value)
    
    def extract_embedded_pdfs(
        self, 
        file_content: bytes, 
        output_dir: Path
    ) -> List[Dict[str, Any]]:
        """
        Extract embedded PDFs from Excel file
        
        Args:
            file_content: Binary content of Excel file
            output_dir: Directory to save extracted PDFs
            
        Returns:
            List of extracted PDF info (filename, path, size)
            
        Note:
            openpyxl has limited support for embedded objects.
            For production, may need to use win32com (Windows) or 
            extract directly from .xlsx ZIP structure.
        """
        extracted_pdfs = []
        
        try:
            workbook = load_workbook(io.BytesIO(file_content))
            
            # Note: This is a placeholder implementation
            # Full implementation would need to:
            # 1. Unzip .xlsx file (it's a ZIP archive)
            # 2. Look for embedded objects in xl/embeddings/
            # 3. Extract and save PDFs
            # 4. Map PDFs to specific cells/controls
            
            # For MVP, we'll rely on validation that PDFs exist
            # and assume external PDF upload if needed
            
            pass  # TODO: Implement full PDF extraction
            
        except Exception as e:
            raise ValueError(f"Failed to extract embedded PDFs: {str(e)}")
        
        return extracted_pdfs
    
    def calculate_completion_stats(self, parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate assessment completion statistics"""
        total_controls = 0
        implemented = 0
        partial = 0
        not_started = 0
        
        for domain in parsed_data.get("domains", []):
            for control in domain.get("controls", []):
                total_controls += 1
                status = control.get("status", "").lower()
                
                if status == "implemented":
                    implemented += 1
                elif status == "partial":
                    partial += 1
                elif status in ["not started", "not_started"]:
                    not_started += 1
        
        completion_percentage = (
            (implemented / total_controls * 100) if total_controls > 0 else 0
        )
        
        return {
            "total_controls": total_controls,
            "implemented": implemented,
            "partial": partial,
            "not_started": not_started,
            "completion_percentage": round(completion_percentage, 2)
        }


# Singleton instance
_excel_processor = None

def get_excel_processor() -> IM8ExcelProcessor:
    """Get or create IM8ExcelProcessor instance"""
    global _excel_processor
    if _excel_processor is None:
        _excel_processor = IM8ExcelProcessor()
    return _excel_processor
